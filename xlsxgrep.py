#!/usr/bin/env python3

import argparse
from openpyxl import load_workbook
import json

parser = argparse.ArgumentParser(description='Find string in xlsx')
parser.add_argument('file', metavar='F', type=str, help='input file')
parser.add_argument('substring', metavar='S', type=str,
        help='substring to find')
parser.add_argument('--first', action='store_true',
        help='Show only first result')
parser.add_argument('--rows', action='store_true',
        help='Show only row numbers')
parser.add_argument('--columns', action='store_true',
        help='Show only column numbers')

args = parser.parse_args()

wb = load_workbook(args.file)
ws = wb.active

rows = ws.rows
for i in range(0, ws.max_row):
    row = rows[i]
    datarow = {}
    for j in range(0, len(row)):
        val = str(row[j].value)
        if val.find(args.substring) >= 0:
            if args.rows:
                print(i)
            elif args.columns:
                print(j)
            else:
                print('{0},{1} {2}'.format(i, j, val))
            if args.first:
                exit(0)
