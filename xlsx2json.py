#!/usr/bin/env python3

import argparse
from openpyxl import load_workbook
import json

parser = argparse.ArgumentParser(description='Convert xlsx to json')
parser.add_argument('file', metavar='F', type=str,
                            help='input file')
parser.add_argument('--header_row', type=int, default=1, help='header row')
parser.add_argument('--first_data_row', type=int, default=2, help='first data row')

args = parser.parse_args()

if args.first_data_row < args.header_row:
    args.first_data_row = args.header_row + 1

wb = load_workbook(args.file)
ws = wb.active

def CellName(x):
    N = ord('Z') - ord('A') + 1
    s = ''
    while True:
        symb = chr(ord('A') + (x % N))
        s = symb + s
        x = x // N
        if x == 0:
            break
        x = x - 1
    return s

header = []
for j in range(1, ws.max_column + 1):
    column_name = ws.cell(row=args.header_row, column=j).value
    if column_name is None:
      column_name = '#' + CellName(j - 1)
    header.append(str(column_name))

data = []
rows = ws.rows
for i in range(args.first_data_row - 1, ws.max_row):
    row = rows[i]
    datarow = {}
    for j in range(0, len(row)):
        datarow[header[j]] = row[j].value
    data.append(datarow)

outdata = {'header': header, 'data': data}

print(json.dumps(outdata, ensure_ascii=False))
